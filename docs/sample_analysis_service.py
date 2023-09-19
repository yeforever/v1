#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@version: 1.0
@author: adair
@contact: adair.ma@amdigital.cn
@time: 2019/09/03 15:52
"""
import os

import pandas as pd
import wordcloud as wc
import jieba
import numpy as np
import re
from PIL import Image

from src.commons.logging_helper import LoggingHelper
from src.config.main import analysis_config
from src.config.main import ttf_config
from openpyxl import Workbook
from matplotlib import pyplot as plt
from matplotlib.ticker import FuncFormatter
from matplotlib.font_manager import FontProperties
from collections import Counter
from pandas import DataFrame
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as Im

from src.config.params_local import remove_words_list
from src.models.mysql.dts_models import AutoBriefTask

APP_NAME = 'xxx_analysis_service'
logger = LoggingHelper.get_logger(APP_NAME)


class XXXAnalysisService(object):
    """
    XXX 平台 数据分析服务
    """
    # todo 将全局可配置参数放在首行，如果配置文件key 值修改，只用在首行修改  例如，config['Deng'] 改成了 config['123'],不需要全文搜索去改
    # todo 提取复用次数较多的和需要特别注释的参数，其余只用1次的可以不用提出来

    # plt画图 固定 字体
    PLT_FONT = FontProperties(fname=ttf_config['simhei'], size=15)

    # wordcloud 固定字体
    WC_FONT = ttf_config['Deng']

    # wordcloud 固定背景图片
    WC_BACKGROUND = ttf_config['ball_png']

    # 需要分析的源文件excel 路径
    SOURCE_EXCEL_PATH = analysis_config['mweibo_cn']

    # 执行分析
    @classmethod
    def run_analysis(cls, task_uuid):
        """
        获取分析图片
        :param task_uuid:
        :return:
        """
        # 断点
        logger.info('==========图表分析开始============')

        # 根据task_uuid 判断 该任务mongodb 有无数据，如无数据，则后续流程无需执行
        result = AutoBriefTask.get_or_none(AutoBriefTask.task_uuid == task_uuid)
        if result:
            sum_task_num = result.sum_task_num
            if sum_task_num is None and sum_task_num == 0:
                logger.info(f'{task_uuid}导出的数据为空！')
                return False

        task = AutoBriefTask.select().where(AutoBriefTask.task_uuid == task_uuid)
        task_keyword = task[0].task_keyword
        try:
            # 制作analysis excel 表
            cls.analysis_excel(task_keyword)
        except Exception as e:
            # 捕捉到异常后输出到日志中
            logger.info(e)
            return False

    # 分析excel
    @classmethod
    def analysis_excel(cls, wordss):
        """
        pandas 读取path 下的excel 文件，制作分析图表
        :param path:文件路径
        :param wordss: 关键词
        :return:
        """
        path = cls.SOURCE_EXCEL_PATH.format(keyword=wordss)
        # 将 源数据excel数据 转换成 dataframe
        df = pd.read_excel(path)

        # analysis 需要 5个sheet
        wb = Workbook()
        # sheet1 改名 分析图
        ws_img = wb.active
        ws_img.title = u'分析图'
        ws_img.append(['名称', '分析图'])

        # sheet2 新建 kol级别分布表 的sheet并插入数据
        wb.create_sheet(u'kol级别分布表')
        ws_kol = wb['kol级别分布表']
        ws_kol.append(['级别名', 'kol人数'])
        # 生成kol级别分布图 并写入kol分级数据
        kol_data = cls.get_group(df, wordss)
        for i in kol_data:
            ws_kol.append(i)
        logger.info('{word}_KOL级别分布数据表,保存成功'.format(word=wordss))
        # 图片插入到sheet1 B2
        ws_img.column_dimensions['B'].width = 60
        ws_img.column_dimensions['A'].width = 60
        ws_img.row_dimensions[2].height = 200
        img = Im(
            analysis_config['mweibo_png']['kol'].format(keyword=wordss))
        img.width, img.height = 450, 250
        ws_img['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws_img['A2'] = 'kol级别分布表'
        ws_img.add_image(img, 'B2')
        logger.info('{word}_KOL级别分布图插入成功'.format(word=wordss))

        # sheet3 新建 kol发文趋势表 的sheet并插入数据
        wb.create_sheet(u'kol发文趋势表')
        ws_time = wb['kol发文趋势表']
        ws_time.append(['时间', '发文数'])
        # 生成kol级别分布图 并写入kol分级数据
        time_data = cls.get_posttime(df, wordss)
        for i in time_data:
            ws_time.append(i)
        logger.info('{word}_kol发文趋势数据表,保存成功'.format(word=wordss))
        # 图片插入到sheet1 B3
        ws_img.row_dimensions[3].height = 200
        img = Im(
            analysis_config['mweibo_png']['post_time'].format(keyword=wordss))
        img.width, img.height = 450, 250
        ws_img['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws_img['A3'] = 'kol发文趋势图'
        ws_img.add_image(img, 'B3')
        logger.info('{word}_kol发文趋势图插入成功'.format(word=wordss))

        # sheet4 新建 kol发文频次表 的sheet并插入数据
        wb.create_sheet(u'kol发文频次表')
        ws_num = wb['kol发文频次表']
        ws_num.append(['账号名称', '发文数'])
        # 生成 kol发文频次图 并写入 kol发文频次 数据
        num_data = cls.get_name(df, wordss)
        for i in num_data:
            ws_num.append(i)
        logger.info('{word}_kol发文频次数据表,保存成功'.format(word=wordss))
        # 图片插入到sheet1 B4
        ws_img.row_dimensions[4].height = 200
        img = Im(
            analysis_config['mweibo_png']['post_article_num'].format(keyword=wordss))
        img.width, img.height = 450, 250
        ws_img['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws_img['A4'] = 'kol发文频次图'
        ws_img.add_image(img, 'B4')
        logger.info('{word}_kol发文频次图插入成功'.format(word=wordss))

        # sheet5 新建 文章内容词频top500 的sheet并插入数据
        wb.create_sheet(u'文章内容词频top500')
        ws_content = wb['文章内容词频top500']
        ws_content.append(['词语', '词频'])
        # 生成 文章内容词频top500词云图 并写入 文章内容词频top500 数据
        content_data = cls.get_content(df, wordss)
        for i in content_data:
            ws_content.append(i)
        logger.info('{word}_文章内容词频top500数据表,保存成功'.format(word=wordss))
        # 图片插入到sheet1 B5
        ws_img.row_dimensions[5].height = 200
        img = Im(
            analysis_config['mweibo_png']['content'].format(keyword=wordss))
        img.width, img.height = 450, 250
        ws_img['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws_img['A5'] = '文章内容top20词频图'
        ws_img.add_image(img, 'B5')
        logger.info('{word}_文章内容top20词频图插入成功'.format(word=wordss))
        # 图片插入到sheet1 B7
        ws_img.row_dimensions[7].height = 200
        img = Im(analysis_config['mweibo_png']['content_wc'].format(keyword=wordss))
        img.width, img.height = 450, 250
        ws_img['A7'].alignment = Alignment(horizontal='center', vertical='center')
        ws_img['A7'] = '文章内容词云top500'
        ws_img.add_image(img, 'B7')
        logger.info('{word}_文章内容词云top500词频图插入成功'.format(word=wordss))

        # sheet7新建 topic内容词频top500 的sheet并插入数据
        wb.create_sheet(u'topic内容词频top500')
        ws_topic = wb['topic内容词频top500']
        ws_topic.append(['词语', '词频'])
        # 生成 topic内容词频top500 并写入 topic内容词频top500 数据
        topic_data = cls.get_topic(df, wordss)
        for i in topic_data:
            ws_topic.append(i)
        logger.info('{word}_topic内容词频top500数据表,保存成功'.format(word=wordss))
        # 图片插入到sheet1 B6
        ws_img.row_dimensions[6].height = 200
        img = Im(
            analysis_config['mweibo_png']['topic'].format(keyword=wordss))
        img.width, img.height = 450, 250
        ws_img['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws_img['A6'] = 'Topic20词频统计图'
        ws_img.add_image(img, 'B6')
        logger.info('{word}_Topic20词频统计图插入成功'.format(word=wordss))
        # 判断是否有topic词云图
        topic_not_none = topic_data
        # 如果有数据
        if topic_not_none:
            # 图片插入到sheet1 B8
            logger.info('{word}_topic内容词频top500词云图插入成功'.format(word=wordss))
            ws_img.row_dimensions[8].height = 200
            img = Im(analysis_config['mweibo_png']['topic_wc'].format(keyword=wordss))
            img.width, img.height = 450, 250
            ws_img['A8'].alignment = Alignment(horizontal='center', vertical='center')
            ws_img['A8'] = 'topic内容词云top500'
            ws_img.add_image(img, 'B8')
            logger.info('{word}_topic内容词云top500插入成功'.format(word=wordss))

        # analysis 文件保存
        wb.save(path.split('.xlsx')[0] + '_analysis.xlsx')
        logger.info('{word}_analysis.xlsx 保存成功'.format(word=wordss))

        # 图片插入完成后删除图片文件
        os.remove(
            analysis_config['mweibo_png']['kol'].format(keyword=wordss))
        os.remove(
            analysis_config['mweibo_png']['post_time'].format(keyword=wordss))
        os.remove(
            analysis_config['mweibo_png']['post_article_num'].format(keyword=wordss))
        os.remove(analysis_config['mweibo_png']['content'].format(keyword=wordss))
        if topic_not_none:
            os.remove(
                analysis_config['mweibo_png']['topic_wc'].format(keyword=wordss))
        os.remove(analysis_config['mweibo_png']['topic'].format(keyword=wordss))
        os.remove(analysis_config['mweibo_png']['content_wc'].format(keyword=wordss))

    # 绘制KOL级别分布折线图
    @classmethod
    def get_group(cls, df, wordss):
        """
        绘制KOL级别分布折线图
        :param wordss:
        :return:
        """
        # 根据 微博链接，账号链接 值 去除重复行
        df = df.drop_duplicates(['微博链接'])
        df = df.drop_duplicates(['账号链接'])

        # 定义级别范围
        bins = [-1, 1000000, 5000000, float('inf')]
        labels = ['底部kol', '腰部kol', '头部kol']

        # 级别分类
        df['kol级别'] = pd.cut(
            df['账号粉丝数'],
            bins,
            right=False,
            labels=labels
        )

        # 账号个数 聚合
        kol_df = df.groupby(df['kol级别']).count()['账号粉丝数'].sort_index(ascending=False)

        # 转换成新的dataframe
        kol_df = pd.DataFrame({'key': kol_df.keys(), 'bad': kol_df})
        df_new = DataFrame()
        df_new['type'] = list(kol_df['key'].values)
        df_new['values'] = list(kol_df['bad'].values)
        df_new['percent'] = [(round(i / sum(df_new['values']), 4)) for i in df_new['values']]

        # 绘制折线图
        _x = df_new['type']
        _y = df_new['percent']
        plt.figure(figsize=(20, 10), dpi=80)
        plt.plot(_x, _y)
        plt.grid(alpha=0.2)

        # 格式化y轴刻度
        def to_percent(temp, position):
            return str(temp) + '%'

        plt.gca().yaxis.set_major_formatter(FuncFormatter(to_percent))

        # 设置x轴
        plt.xticks(range(len(_x)), rotation=45, fontproperties=cls.PLT_FONT)
        for a, b in zip(list(range(len(_x))), list(_y)):
            plt.text(a, b, '{:.2f}%'.format(b * 100), ha='center', va='bottom', fontsize=14)
        plt.xlabel('kol级别', fontproperties=cls.PLT_FONT)
        plt.ylabel('kol占比', fontproperties=cls.PLT_FONT)
        plt.title('kol占比折线图', fontproperties=cls.PLT_FONT)
        plt.savefig(analysis_config['mweibo_png']['kol'].format(
            keyword=wordss))
        logger.info('{word}_KOL级别分布折线图,保存成功'.format(word=wordss))
        plt.close()

        # kol级别分布数据 作为返回值返回
        kol_data = []
        kol_data.extend(list(i) for i in kol_df.values)
        return kol_data

    # 绘制KOL发文趋势折线图
    @classmethod
    def get_posttime(cls, df, wordss):
        """
        绘制KOL发文趋势折线图
        :param wordss:
        :return:
        """
        # 去除重复行
        df = df.drop_duplicates(['微博链接'])

        # 获得df的行数
        height = df.shape[0]
        df['发布时间'] = pd.to_datetime(df['发布时间'])
        df = df.sort_values(by='发布时间')
        df.set_index('发布时间', inplace=True)
        df['发文数量'] = [1 for x in range(height)]

        # 重采样， 'M' 按 月份统计频率
        df = df.resample('M').sum()

        # 绘制折线图
        _x = df.index
        _x = [i.strftime('%Y-%m') for i in _x]
        _y = df['发文数量']
        plt.figure(figsize=(20, 10), dpi=80)
        plt.grid(alpha=0.2)
        plt.plot(_x, _y)
        plt.xticks(range(len(_x)), list(_x), rotation=45, fontproperties=cls.PLT_FONT)
        plt.yticks(fontproperties=cls.PLT_FONT)
        plt.xlabel('时间', fontproperties=cls.PLT_FONT)
        plt.ylabel('发文数量', fontproperties=cls.PLT_FONT)
        plt.title('发文趋势折线图', fontproperties=cls.PLT_FONT)

        for a, b in zip(list(range(len(_x))), list(_y)):
            plt.text(a, b, b, ha='center', va='bottom', fontsize=14)
        plt.savefig(analysis_config['mweibo_png']['post_time'].format(
            keyword=wordss))
        logger.info('{word}_KOL发文趋势折线图,保存成功'.format(word=wordss))
        plt.close()

        # KOL发文趋势数据 作为返回值返回
        kol_data = []
        kol_data.extend([list(_x)[i], list(_y)[i]] for i in range(len(list(_x))))
        return kol_data

    # 绘制KOL发文频率折线图
    @classmethod
    def get_name(cls, df, wordss):
        """
        绘制KOL发文频率折线图
        :param wordss:
        :return:
        """
        df = df.drop_duplicates(['微博链接'])

        size = df.shape[0]
        df['发文数量'] = [1 for i in range(size)]

        df.set_index(df['账号名称'], inplace=True)
        grouped = df['发文数量'].groupby(df['账号名称'])
        df['post_times'] = grouped.sum()

        df = df.drop_duplicates(['账号名称'])

        df_kol = df.sort_values(by='post_times', ascending=False)
        # 选取行列标号索引
        df_new = df_kol.iloc[0:20, :]
        df_new.set_index(df_new['账号名称'], inplace=True)

        _x = [i[:len(i) // 2] + '\n' + i[len(i) // 2:] if len(i) >= 6 else i for i in df_new['账号名称']]
        _y = df_new['post_times']
        plt.figure(figsize=(25, 14), dpi=80)
        plt.plot(_x, _y)
        plt.grid(alpha=0.2)
        plt.xticks(range(len(_x)), list(_x), rotation=45, fontproperties=cls.PLT_FONT)
        plt.yticks(fontproperties=cls.PLT_FONT)
        for a, b in zip(list(range(len(_x))), list(_y)):
            plt.text(a, b, b, ha='center', va='bottom', fontsize=14)
        plt.xlabel('用户名', fontproperties=cls.PLT_FONT)
        plt.ylabel('发文数量', fontproperties=cls.PLT_FONT)
        plt.title('发文数量折线图', fontproperties=cls.PLT_FONT)
        plt.savefig(analysis_config['mweibo_png']['post_article_num'].format(
            keyword=wordss))
        logger.info('{word}_KOL发文频率折线图,保存成功'.format(word=wordss))
        plt.close()

        num_data = []
        num_data.extend([list(_x)[i], list(_y)[i]] for i in range(len(list(_x))))
        return num_data

    # 绘制文章内容TOP20词频统计折线图
    @classmethod
    def get_content(cls, df, wordss):
        """
        绘制文章内容TOP20词频统计折线图
        :param wordss:
        :return:
        """
        # 获取所有内容 并切词
        article = list(df['微博内容'])
        article = ''.join(article)
        article_list = jieba.lcut(article)
        real_word = []
        for word in article_list:
            if word not in remove_words_list:
                real_word.append(word)
        word_count = Counter(real_word)

        # 绘制词云图
        cls.get_content_wc(word_count,wordss)

        # top 20 和 top 500
        content_top_20 = word_count.most_common(20)
        content_top_500 = word_count.most_common(500)
        df_word_20 = DataFrame()
        df_word_500 = DataFrame()

        df_word_20['高频词语'] = [i[0][:len(i[0]) // 2] + '\n' + i[0][len(i[0]) // 2:] if len(i[0]) >= 6 else i[0] for i
                              in content_top_20]
        df_word_20['出现次数'] = [i[1] for i in content_top_20]

        df_word_500['高频词语'] = [i[0] for i in content_top_500]
        df_word_500['出现次数'] = [i[1] for i in content_top_500]

        # top 20 绘制词频统计折线图
        plt.figure(figsize=(20, 10), dpi=80)
        _x = df_word_20['高频词语']
        _y = df_word_20['出现次数']
        plt.plot(_x, _y)
        plt.xticks(range(len(_x)), list(_x), rotation=45, fontproperties=cls.PLT_FONT)
        plt.yticks(fontproperties=cls.PLT_FONT)
        for a, b in zip(list(range(len(_x))), list(_y)):
            plt.text(a, b, b, ha='center', va='bottom', fontsize=15)
        plt.grid(alpha=0.2)
        plt.xlabel('高频词语', fontproperties=cls.PLT_FONT)
        plt.ylabel('出现次数', fontproperties=cls.PLT_FONT)
        plt.title('正文词频统计', fontproperties=cls.PLT_FONT)
        plt.savefig(analysis_config['mweibo_png']['content'].format(
            keyword=wordss))
        file = analysis_config['mweibo_png']['content'].format(
            keyword=wordss)
        plt.savefig(file)
        logger.info('mweibo_{word}_文章内容TOP20词频统计折线图,保存成功'.format(word=wordss))
        plt.close()

        # top 500 写入 sheet 文章内容词频top500
        _x1 = df_word_500['高频词语']
        _y1 = df_word_500['出现次数']
        content_data = []
        content_data.extend([list(_x1)[i], list(_y1)[i]] for i in range(len(list(_x1))))
        return content_data

    # 绘制Topic20词频统计折线图
    @classmethod
    def get_topic(cls, df, wordss):
        """
        绘制Topic20词频统计折线图
        :param wordss:
        :return:
        """
        df = df.drop_duplicates(['微博链接'])

        article_list = list(df['微博内容'])

        all_article = []
        for i in article_list:
            article = re.findall(r'#(.*?)#', i, re.S)
            all_article.append(article)
        all_article = [i[0] for i in all_article if i != []]
        if not all_article:
            return []

        all_article = []
        for i in article_list:
            article = re.findall(r'#(.*?)#', i, re.S)
            all_article.append(article)
        all_article = [i[0] for i in all_article if i != []]
        all_article = ''.join(all_article)
        all_article = jieba.lcut(all_article)
        real_word_list = []
        for word in all_article:
            if word not in remove_words_list:
                real_word_list.append(word)
        topic_word = Counter(real_word_list)

        # 绘制词云图
        cls.get_topic_wc(topic_word, wordss)

        topic_top_20 = topic_word.most_common(20)
        topic_top_500 = topic_word.most_common(500)
        df_word_20 = DataFrame()
        df_word_500 = DataFrame()
        df_word_20['Topic词语'] = [i[0][:len(i[0]) // 2] + '\n' + i[0][len(i[0]) // 2:] if len(i[0]) >= 8 else i[0]
                                 for i in topic_top_20]
        df_word_20['出现次数'] = [i[1] for i in topic_top_20]
        df_word_500['Topic词语'] = [i[0] for i in topic_top_500]
        df_word_500['出现次数'] = [i[1] for i in topic_top_500]

        _x = df_word_20['Topic词语']
        _y = df_word_20['出现次数']
        plt.figure(figsize=(20, 12), dpi=80)
        plt.plot(_x, _y)
        plt.xticks(range(len(_x)), list(_x), rotation=45, fontproperties=cls.PLT_FONT)
        plt.yticks(fontproperties=cls.PLT_FONT)
        plt.grid(alpha=0.2)
        for a, b in zip(list(range(len(_x))), list(_y)):
            plt.text(a, b, b, ha='center', va='bottom', fontsize=14)
        plt.xlabel('Topic高频词语', fontproperties=cls.PLT_FONT)
        plt.ylabel('出现次数', fontproperties=cls.PLT_FONT)
        plt.title('Topic词频统计', fontproperties=cls.PLT_FONT)
        plt.savefig(analysis_config['mweibo_png']['topic'].format(
            keyword=wordss))
        file = analysis_config['mweibo_png']['topic'].format(
            keyword=wordss)
        plt.savefig(file)
        logger.info('mweibo_{word}_Topic20词频统计折线图,保存成功'.format(word=wordss))
        plt.close()

        _x1 = df_word_500['Topic词语']
        _y1 = df_word_500['出现次数']
        topic_data = []
        topic_data.extend([list(_x1)[i], list(_y1)[i]] for i in range(len(list(_x1))))
        return topic_data

    # 绘制正文词云图
    @classmethod
    def get_content_wc(cls, word_count, wordss):
        """
        绘制正文词云图
        :param word_count:
        :return:
        """
        # 绘制词云图
        mask = np.array(Image.open(cls.WC_BACKGROUND))
        pic_config = wc.WordCloud(
            font_path=cls.WC_FONT,
            max_words=500,
            mask=mask,
            max_font_size=180,
            min_font_size=1,
            background_color='white'
        )
        pic_config.generate_from_frequencies(word_count)
        pic_config.to_file(
            analysis_config['mweibo_png']['content_wc'].format(keyword=wordss))
        logger.info('mweibo_{word}_正文词云图,保存成功'.format(word=wordss))

    # 绘制topic词云图
    @classmethod
    def get_topic_wc(cls, topic_word, wordss):
        """
        绘制topic词云图
        :param topic_word:
        :return:
        """
        mask = np.array(Image.open(cls.WC_BACKGROUND))
        pic_config = wc.WordCloud(
            font_path=cls.WC_FONT,
            max_words=500,
            mask=mask,
            max_font_size=180,
            min_font_size=1,
            background_color='white'
        )
        pic_config.generate_from_frequencies(topic_word)
        pic_config.to_file(
            analysis_config['mweibo_png']['topic_wc'].format(keyword=wordss))
        logger.info('mweibo_{word}_topic词云图,保存成功'.format(word=wordss))


if __name__ == '__main__':
    task_uuid = '3333ac2e-14e7-4fde-a193-abbcc4aa805b'
    XXXAnalysisService.run_analysis(task_uuid=task_uuid)
