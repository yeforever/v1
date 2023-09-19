"""
@version: 1.0
@author: anne
@contact: thy.self@foxmail.com
@time: 2021/11/30 下午3:24
"""
import json
import os
import re
import time
import uuid
import zipfile
from collections import Counter

import jieba
import openpyxl
import requests
from PIL import Image
from openpyxl.drawing.image import Image as Im
from loguru import logger
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.font_manager import FontProperties
from openpyxl.styles import Alignment
from pandas.core.frame import DataFrame
import pandas as pd
from peewee import fn
from wordcloud import WordCloud

from src.commons.common_job_models_alias import wx, wb, rb, bi, dy, job, task, stage
from src.commons.common_job_params import platform_code_worker, platform_code_pages, \
    platform_code_query, platform_code_wc_field, platform_code_zh, pymysql_conn, platform_code_excel_df_title, \
    platform_code_wc_df_title, simhei, ciyun, deng, ball, platform_code_en
from src.commons.datatime_helper import DateTimeHelper
from src.config.main import system_config
from src.config.params_local import remove_words_list

# matplotlib.use('Agg') 服务端文件导出出错,加上这句导出不报错
matplotlib.use('Agg')

# pandas 直接通过conn 读取mysql 数据
conn = pymysql_conn

# matplotlib 绘制图形图表配置字体
PLT_FONT = FontProperties(fname=simhei, size=15)


class JobServer(object):
    @classmethod
    def submit_job_to_mysql(cls, job_uuid, job_desc, job_name, remarks, create_user_name):
        job_created_time = DateTimeHelper.format_datetime(int(time.time()))
        job_code = DateTimeHelper.format_datetime(int(time.time()), '%Y%m%d')
        is_job_time_exist = job.get_or_none(job.job_code.contains(job_code))
        if is_job_time_exist:
            job_code = job.select(job.job_code).where(
                job.job_code.contains(job_code))
            job_code_list = [i.job_code for i in job_code]
            max_job_code = max(job_code_list)
            now_job_code = str(int(max_job_code) + 1)
        else:
            now_job_code = f'{job_code}001'
        job.create(
            job_code=now_job_code,
            job_created_time=job_created_time,
            job_run_time=job_created_time,
            job_finish_time=job_created_time,
            job_desc=json.loads(job_desc),
            job_name=job_name,
            job_status=2,
            job_uuid=job_uuid,
            remarks=remarks,
            create_user_name=create_user_name
        )
        # TODO 推送1，提交
        cls.send_51wom_markdown_information(create_user_name, now_job_code, job_created_time, 0, '执行中',
                                            DateTimeHelper.format_datetime(int(time.time()) + 120))
        job_content = json.loads(job_desc)['job_content']
        for index, i in enumerate(job_content):
            stage_uuid = str(uuid.uuid4())
            platforms = i['platform'].split('/')
            keywords = i['keywords'].split('/')
            keyword_temp = ''.join(keywords)
            download_url = f'https://brief.womdata.com/dts-auto-brief/download?stage_uuid={stage_uuid}&' \
                           f'file_name={keyword_temp}'
            logger.info(download_url)
            stage.create(
                stage_uuid=stage_uuid,
                job_uuid=job_uuid,
                stage_name=job_name + f'_{index}',
                stage_desc=str(i),
                stage_status=2,
                stage_created_time=job_created_time,
                stage_run_time=job_created_time,
                stage_finish_time=job_created_time,
                download_url=download_url,
                limit=i['limitnum']
            )

            platforms = list(map(lambda x: int(x), list(set(platforms))))
            # fixme 此处3层 for 循环，需要优化
            for platform in platforms:
                for keyword in list(set(keywords)):
                    task_uuid = str(uuid.uuid4())
                    task_limit_starttime = DateTimeHelper.format_datetime(
                        DateTimeHelper.parse_formatted_datetime(str(i['starttime']), '%Y%m%d'))
                    task_limit_endtime = DateTimeHelper.format_datetime(
                        DateTimeHelper.parse_formatted_datetime(str(i['endtime']), '%Y%m%d'))
                    task_limit_num = i['limitnum']
                    task.create(
                        job_uuid=job_uuid,
                        task_uuid=task_uuid,
                        stage_uuid=stage_uuid,
                        task_status=2,
                        task_run_starttime=job_created_time,
                        task_run_endtime=job_created_time,
                        platform_code=platform,
                        task_desc={platform_code_zh[platform]: keyword},
                        task_keyword=keyword,
                        task_limit_num=task_limit_num,
                        task_limit_starttime=task_limit_starttime,
                        task_limit_endtime=task_limit_endtime
                    )
                    cls.run_task(stage_uuid, task_uuid, platform, keyword, task_limit_num,
                                 task_limit_starttime, task_limit_endtime)
            # todo 延时方式打包 stage 发送任务完成，导出开始等待任务结束开始导出打包
            cls.export_stage(stage_uuid, 2)
        # TODO 推送2，完成
        cls.send_reply_text(create_user_name)
        logger.info(('============ 校验工单任务 ===========  \n'
                     'job_uuid: {job_uuid}, '
                     'job_name: {job_name}, '
                     'job_content: {job_content},').format(
            job_uuid=job_uuid,
            job_name=job_name,
            job_content=str(job_content)
        ))

    @classmethod
    def run_task(cls, stage_uuid, task_uuid, platform, keyword, task_limit_num, task_limit_starttime,
                 task_limit_endtime):
        logger.info('=================== task 发送celery =====================')
        app = platform_code_worker[platform]
        pages = int(task_limit_num / platform_code_pages[platform])
        for i in range(1, pages + 1):
            task_info = {
                'task_uuid': task_uuid,
                'stage_uuid': stage_uuid,
                'start_ts': DateTimeHelper.parse_formatted_datetime(task_limit_starttime, '%Y-%m-%d %H:%M:%S'),
                'end_ts': DateTimeHelper.parse_formatted_datetime(task_limit_endtime, '%Y-%m-%d %H:%M:%S') + 3600,
                'keyword': keyword,
                'page': i
            }
            app.send_task(f'{platform_code_en[platform]}.{platform_code_en[platform]}', [task_info])
        logger.info(f'task_uuid: {task_uuid} 发送成功')

    @classmethod
    def export_stage(cls, stage_uuid, export_type):
        """

        :param stage_uuid:
        :param export_type: 1: 不延时，2： 延时
        :return:
        """
        delay_sec = 30 if export_type == 1 else 0
        time.sleep(delay_sec)
        query = task.select(
            task.id,
            task.task_uuid,
            task.platform_code,
            task.task_keyword,
            task.task_limit_starttime,
            task.task_limit_endtime,
        ).where(
            task.stage_uuid == stage_uuid
        )
        stage_path = f"{system_config['path']}/data/{stage_uuid}"

        for i in query:
            # 每个task 给10s 时间 延时来等待采集完成
            delay_sec = 10 if export_type == 2 else 0
            time.sleep(delay_sec)
            task_id = i.id
            task_uuid = i.task_uuid
            platform_code = i.platform_code
            if i.platform_code == 4:
                # B站链路长
                time.sleep(delay_sec)
            if i.platform_code == 3:
                # 小红书等待时间长一点，1分钟
                time.sleep(delay_sec * 6)
            task_keyword = i.task_keyword
            start = str(i.task_limit_starttime).split(' ')[0].replace('-', '')
            end = str(i.task_limit_endtime).split(' ')[0].replace('-', '')

            source_df = cls.export_task_source(task_uuid, platform_code)
            if not os.path.exists(stage_path):
                os.makedirs(stage_path)
            source_df.to_excel(
                f'{stage_path}/{task_keyword}_{platform_code_zh[platform_code]}_{start}_{end}_{task_id}.xlsx')

            cls.export_task_analysis(
                task_uuid, platform_code, source_df,
                f'{stage_path}/{task_keyword}_{platform_code_zh[platform_code]}_{start}_{end}_{task_id}_analysis.xlsx')
        # 生成报告页面词云图
        platforms = list(set(x.platform_code for x in query))
        for i in platforms:
            pl = platform_code_query[i]
            cont_query = pl.select(
                platform_code_wc_field[i][0][0],
                platform_code_wc_field[i][0][1],
            ).where(
                pl.stage_uuid == stage_uuid
            )
            article_text = ''.join(x.__data__.get(platform_code_wc_field[i][1][0]) for x in list(cont_query))
            article_text = repr(''.join(
                str(article_text).replace("请问", "").replace("\n", "").replace("\t", "").replace("\r", "").replace("捂脸",
                                                                                                                  "").
                replace("抖音小助手", "").replace("抖音", "").replace("助手", "").replace("助理", "").replace("发货地址", "")))
            topic_text = ''.join(x.__data__.get(platform_code_wc_field[i][1][1]) for x in list(cont_query))
            article_counter = cls.cut_word(article_text)
            topic_counter = cls.cut_topic(topic_text) if i in [2, 5] else cls.cut_word(topic_text)

            mask = np.array(Image.open(ciyun))
            wc = WordCloud(font_path=deng, background_color="white", mask=mask, max_words=1000000)

            if article_counter:
                wc.generate_from_frequencies(dict(article_counter))
                wc.to_file(f'{stage_path}/wc_{str(i)}_1.png')
            if topic_counter:
                wc.generate_from_frequencies(dict(topic_counter))
                wc.to_file(f'{stage_path}/wc_{str(i)}_2.png')

        cls.zip_dir(f'{stage_path}', f"{stage_path}.zip")
        logger.info('打包成功')
        # todo 企微推送通知

    @classmethod
    def cut_word(cls, text):
        return Counter([k for k in jieba.lcut(text, cut_all=False) if k not in remove_words_list]).most_common(500)

    @classmethod
    def cut_topic(cls, text):
        text = re.sub(r'\\[uU][0-9a-z]+', '', text)
        return Counter([k for k in jieba.lcut(''.join(re.findall(r'#(.*?)#| ', text, re.S)), cut_all=False)]
                       ).most_common(500)

    @classmethod
    def export_task_source(cls, task_uuid, platform_code):
        """
        导出task 源数据
        :param platform_code:
        :param task_uuid:
        :return:
        """
        wx_query = wx.select(
            wx.title,
            wx.content,
            wx.post_time,
            wx.url,
            wx.article_pos,
            wx.read,
            wx.like,
            wx.comment,
            wx.user_name,
            wx.user_id,
            wx.user_fans,
            wx.interaction
        ).where(
            wx.task_uuid == task_uuid
        ).group_by(wx.url)
        wb_query = wb.select(
            fn.CONCAT('https://m.weibo.cn/profile/', wb.user_id).alias('user_url'),
            wb.user_id,
            wb.user_name,
            wb.user_fans,
            wb.user_follow,
            wb.user_post,
            wb.url.alias('note_url'),
            wb.content,
            wb.post_time,
            wb.share,
            wb.comment,
            wb.like,
            (wb.like + wb.share + wb.comment).alias('total_interaction'),
            wb.is_repost
        ).where(
            wb.task_uuid == task_uuid
        ).group_by(wb.url)

        rb_query = rb.select(
            rb.user_url,
            rb.user_name,
            rb.user_brief,
            rb.user_level,
            rb.user_location,
            rb.user_fans,
            rb.user_fans_level,
            rb.user_like_num,
            rb.user_follow_num,
            rb.user_collected_num,
            rb.user_collect_like_num,
            rb.user_all_note_num,
            rb.url,
            rb.note_type,
            rb.title,
            rb.content,
            rb.post_time,
            rb.like,
            rb.comment,
            rb.collect,
            rb.share,
            rb.note_collect_like_num,
            rb.interaction,
            rb.score,
            rb.note_cooperate_binds,
            rb.note_tags
        ).where(
            rb.task_uuid == task_uuid
        ).group_by(rb.url)

        bi_query = bi.select(
            bi.video_type,
            bi.post_type,
            bi.url,
            bi.title,
            bi.content,
            bi.img_url,
            bi.play,
            bi.danmu,
            bi.collect,
            bi.comment,
            bi.coin,
            bi.like,
            bi.share,
            bi.interaction,
            bi.tag,
            bi.duration,
            bi.post_time,
            bi.user_name,
            bi.user_id,
            bi.user_url,
            bi.user_follows,
            bi.user_fans
        ).where(
            bi.task_uuid == task_uuid
        ).group_by(bi.url)
        dy_query = dy.select(
            dy.title,
            dy.url,
            dy.content,
            dy.like,
            dy.download,
            dy.share,
            dy.forward,
            dy.comment,
            dy.post_time,
            dy.user_id,
            dy.user_name,
            dy.user_sign,
            dy.user_url,
            dy.user_fans,
            dy.user_follows,
            dy.user_likes,
            dy.interaction
        ).where(
            dy.task_uuid == task_uuid
        ).group_by(dy.url)

        sql_dict = {
            1: str(wx_query),
            2: str(wb_query),
            3: str(rb_query),
            4: str(bi_query),
            5: str(dy_query)
        }
        # conn 长链接会报错 pymysql.err.InterfaceError: (0, '') 每次读之前，ping 一下断开就重连
        # 数据库之所以断开连接，是因为数据库默认的wait_timeout = 28800, 这个单位是秒，换算后是8小时，也就是原来我的服务启动8小时后，就会被mysql自动断开
        conn.ping(reconnect=True)
        df = pd.read_sql(sql_dict[platform_code], con=conn)
        df.columns = platform_code_excel_df_title[platform_code]

        return df

    @classmethod
    def export_task_analysis(cls, task_uuid, platform_code, source_df, file_path):
        """
        导出task 分析数据
        :param platform_code:
        :param task_uuid:
        :param source_df:
        :param file_path:
        :return:
        """
        try:
            file_dir = file_path.replace('_analysis.xlsx', '')

            def plot_draw(df, x, y, title, file_name):
                _x = df[x]
                _y = df[y]
                plt.figure(figsize=(20, 10), dpi=80)
                plt.xticks(range(len(_x)), _x, rotation=45, fontproperties=PLT_FONT)
                for a, b in zip(range(len(_x)), _y):
                    plt.text(a, b, b, ha='center', va='bottom', fontsize=14)
                plt.grid(alpha=0.2)
                plt.plot(_y)
                plt.xlabel(x, fontproperties=PLT_FONT)
                plt.ylabel(y, fontproperties=PLT_FONT)
                plt.title(title, fontproperties=PLT_FONT)

                plt.savefig(file_name)

            # kol级别分布表
            pl = platform_code_query[platform_code]
            # 无数据就不做操作了
            if pl.get_or_none(task_uuid=task_uuid) is None:
                return None

            query = pl.select(
                pl.user_fans_level,
                fn.COUNT(pl.user_name.distinct()).alias('num')
            ).where(
                pl.task_uuid == task_uuid
            ).group_by(
                pl.user_fans_level
            )
            level_df = pd.read_sql(str(query), con=conn)
            level_df.columns = ['级别名', 'kol人数']
            plot_draw(level_df, '级别名', 'kol人数', 'kol级别分布表', f'{file_dir}_3.png')

            # kol发文趋势表
            post_month = fn.YEAR(pl.post_time) * 100 + fn.MONTH(pl.post_time)
            query = pl.select(
                post_month,
                fn.COUNT(pl.id)
            ).where(
                pl.task_uuid == task_uuid
            ).group_by(
                post_month
            )
            month_df = pd.read_sql(str(query), con=conn)
            month_df.columns = ['时间', '发文数']
            plot_draw(month_df, '时间', '发文数', 'kol发文趋势表', f'{file_dir}_4.png')

            # kol发文频次表
            query = pl.select(
                pl.user_name,
                fn.COUNT(pl.id)
            ).where(
                pl.task_uuid == task_uuid
            ).group_by(
                pl.user_name
            )
            name_df = pd.read_sql(str(query), con=conn)
            name_df.columns = ['账号名称', '发文数']
            plot_draw(name_df, '账号名称', '发文数', 'kol发文频次表', f'{file_dir}_5.png')

            # 文章内容词频top500
            # topic内容词频top500

            article_df_name = platform_code_wc_df_title[platform_code][0]
            topic_df_name = platform_code_wc_df_title[platform_code][1]

            article_contents = ''.join(str(i) for i in source_df[article_df_name].values)
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            article_contents = ILLEGAL_CHARACTERS_RE.sub(r'', article_contents)
            topic_contents = ''.join(str(i) for i in source_df[topic_df_name].values)

            article_counter = cls.cut_word(article_contents)
            topic_counter = cls.cut_topic(topic_contents) if platform_code in [2] else cls.cut_word(topic_contents)

            mask = np.array(Image.open(ball))
            wc = WordCloud(font_path=deng, background_color="white", mask=mask, max_words=1000000)

            # 词频df
            # print(article_counter)
            article_df = DataFrame(article_counter)
            article_df.columns = ['词语', '词频']
            # print(topic_counter)
            topic_df = DataFrame(topic_counter)
            topic_df.columns = ['词语', '词频']

            wc.generate_from_frequencies(dict(article_counter))
            wc.to_file(f'{file_dir}_6.png')

            wc.generate_from_frequencies(dict(topic_counter))
            wc.to_file(f'{file_dir}_7.png')

            # excel 插入
            with pd.ExcelWriter(file_path) as writer:
                level_df.to_excel(writer, encoding='utf-8', sheet_name='kol级别分布表', index=False)
                month_df.to_excel(writer, encoding='utf-8', sheet_name='kol发文趋势表', index=False)
                name_df.to_excel(writer, encoding='utf-8', sheet_name='kol发文频次表', index=False)
                article_df.to_excel(writer, encoding='utf-8', sheet_name='文章内容词频top500', index=False)
                topic_df.to_excel(writer, encoding='utf-8', sheet_name='标签内容词频表', index=False)

            workbook = openpyxl.load_workbook(file_path)
            workbook.create_sheet(u'分析图', 0)
            ws = workbook['分析图']
            ws.append(['名称', '分析图'])

            # 分析图
            def insert_img(x_k, y_k, x_v, y_v):
                ws.column_dimensions['B'].width = 60
                ws.column_dimensions['A'].width = 60
                ws.row_dimensions[int(x_k[1])].height = 200
                img = Im(y_v)
                img.width, img.height = 450, 250
                ws[x_k].alignment = Alignment(horizontal='center', vertical='center')
                ws[x_k] = x_v
                ws.add_image(img, y_k)

            insert_img('A2', 'B2', 'kol级别分布表', f'{file_dir}_3.png')
            insert_img('A3', 'B3', 'kol发文趋势表', f'{file_dir}_4.png')
            insert_img('A4', 'B4', 'kol发文频次图', f'{file_dir}_5.png')
            insert_img('A5', 'B5', '内容词云top500', f'{file_dir}_6.png')
            insert_img('A6', 'B6', 'topic词云top500', f'{file_dir}_7.png')
            workbook.save(file_path)
        except Exception as e:
            logger.info('anlysis 导出失败')
            logger.info(e)

    @classmethod
    def zip_dir(cls, input_path, output_path):
        """
        压缩指定文件夹
        :param input_path: 目标文件夹路径
        :param output_path: 压缩文件保存路径+xxxx.zip
        :return: 无
        """
        zip_obj = zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED)
        for path, dirnames, filenames in os.walk(input_path):
            # 去掉目标跟路径，只对目标文件夹下边的文件及文件夹进行压缩
            fpath = path.replace(input_path, '')
            for filename in filenames:
                # 只打包xlsx 文件
                if filename.endswith(".xlsx"):
                    zip_obj.write(os.path.join(path, filename), os.path.join(fpath, filename))
        zip_obj.close()

    @classmethod
    def send_reply_text(cls, created_name):
        try:
            resp = requests.post(
                url="https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=4438c2d5-01f7-4b89-b7b5-d8157935844f",
                headers={"Content-Type": "application/json"},
                data=json.dumps({
                    "msgtype": "text",
                    "text": {
                        "mentioned_list": [created_name],
                        "content": f"亲爱的下单人 {created_name} 您好，51wom采集系统已完成工单采集，请前往51wom下载报告。"
                    }
                }))
            print(resp.text)
            return True
        except Exception as e:
            # logger.error(e)
            return False

    @classmethod
    def send_51wom_markdown_information(cls, created_name, code, created_time, delay_num, status, end_time):
        # select_url = 'http://47.104.170.32:8182/find_select_text'
        try:
            resp = requests.post(
                url="https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=4438c2d5-01f7-4b89-b7b5-d8157935844f",
                headers={"Content-Type": "application/json"},
                data=json.dumps({
                    "msgtype": "markdown",
                    "markdown": {
                        "content": f"【51wom 自动化采集工单】\n"
                        # f" 51wom接收到下单人为 <font color=\"info\">{created_name}</font> 的工单，工单信息如下:\n"
                                   f"> 工单编号:  <font color=\"info\">{code}</font>\n"
                                   f"> 下单人:  <font color=\"info\">{created_name}</font>\n"
                                   f"> 下单时间:  <font color=\"info\">{created_time}</font>\n"
                                   f"> 排队数:  <font color=\"info\">{delay_num}</font>\n"
                                   f"> 工单状态:  <font color=\"info\">{status}</font>\n"
                                   f"> 预计完成时间:  <font color=\"info\">{end_time}</font>"
                        # f"51wom采集系统正在为您处理，请耐心等待,点击上方查询入口，随时掌握工单进度动态。"
                    }
                }))
            print(resp.text)
            return True
        except Exception as e:
            # logger.error(e)
            return False




# if __name__ == '__main__':
#     dy_query = dy.select(
#         dy.title,
#         dy.url,
#         dy.content,
#         dy.like,
#         dy.download,
#         dy.share,
#         dy.forward,
#         dy.comment,
#         dy.post_time,
#         dy.user_id,
#         dy.user_name,
#         dy.user_sign,
#         dy.user_url,
#         dy.user_fans,
#         dy.user_follows,
#         dy.user_likes,
#         dy.interaction
#     ).where(
#         dy.stage_uuid == 'task_uuid'
#     )
#
#     sql_dict = {
#         1: str(wx_query),
#         2: str(wb_query),
#         3: str(rb_query),
#         4: str(bi_query),
#         5: str(dy_query)
#     }
#     # conn 长链接会报错 pymysql.err.InterfaceError: (0, '') 每次读之前，ping 一下断开就重连
#     # 数据库之所以断开连接，是因为数据库默认的wait_timeout = 28800, 这个单位是秒，换算后是8小时，也就是原来我的服务启动8小时后，就会被mysql自动断开
#     conn.ping(reconnect=True)
#     df = pd.read_sql(sql_dict[platform_code], con=conn)
#     df.columns = platform_code_excel_df_title[platform_code]